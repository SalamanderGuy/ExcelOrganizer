
from openpyxl import Workbook, load_workbook
import os
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import datetime
from datetime import timedelta, date
import ErrorWindow
from pathlib import Path


def ExportFile(dates=[""], input_text = ""):
    global filelist, x, item, count, date, value, filename
    filelist = []
    check_input = (input_text != "")

    for file in Path('.').rglob('*.xlsx'):
        x = str(file)
        if not x.startswith("Αναζητησεις Δρομολογίων"):
            filename = x.split(" ", 1)
            date = filename[0]
            itinerary = filename[1]            
            idx = date.rfind('\\')
            if idx != -1:
                date = date[idx + 1:]
            itinerary_date = datetime.datetime.strptime(date, "%d-%m-%Y")
            start = datetime.datetime.strptime(dates[0], "%d-%m-%Y")
            end = datetime.datetime.strptime(dates[-1], "%d-%m-%Y")            
            if start <= itinerary_date <= end:
                if check_input and itinerary.find(input_text) == -1:
                    continue
                filelist.append(x)
    #print(filelist[0].split(" ")[0],filelist[-1].split(" ")[0])

    if not filelist:
        print("Δεν υπάρχουν δρομολόγια με αυτές τις ημερομηνιες")
        return ErrorWindow.on_click()
    list = []
    counter_list = []
    for item in filelist:
        wb = load_workbook(item)
        ws = wb.active

        count = 0
        i = 6
        onoma = ""
        while onoma != None:
            onoma = ws["C" + str(i)].value
            klados = ws["D" + str(i)].value
            bathmos = ws["E" + str(i)].value
            parousia1 = ws["F" + str(i)].value
            parousia2 = ws["G" + str(i)].value

            if onoma == None:
                counter_list.append(count)
                break

            if i == 6:
                date = item.split(" ")[0]
                idx = date.rfind('\\')
                if (idx != -1):
                    date = date[idx + 1:]
                it = item.split(" ")[1].split(".")[0]
                list.append((date, it, onoma, klados, bathmos, parousia1, parousia2))
            else:
                list.append((None, None, onoma, klados, bathmos, parousia1, parousia2))
            i = i + 1
            count += 1
    #print(list)
    #print(counter_list)
    #print(len(list[0]))
    wb = Workbook()
    ws = wb.active
    # print(type(ws))
    ws.append(["Ημερομηνία", "Δρομολόγιο", "Επιβάτης", "Κλάδος", "Βαθμός", "Παρουσία_1", "Παρουσία_2"])
    for i in list:
        #print(i)
        ws.append(i)
    x = 2
    for value in counter_list:
        ws.merge_cells("A" + str(x) + ":A" + str(x + value - 1))
        ws["A" + str(x)].alignment = Alignment(horizontal="center")
        ws["A" + str(x)].alignment = Alignment(vertical="center")
        ws.merge_cells("B" + str(x) + ":B" + str(x + value - 1))
        ws["B" + str(x)].alignment = Alignment(horizontal="center")
        ws["B" + str(x)].alignment = Alignment(vertical="center")
        x += value
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(wrap_text=True)
    ws['B1'].font = Font(bold=True, size=13)
    ws['C1'].font = Font(bold=True, size=13)
    ws['D1'].font = Font(bold=True, size=13)
    ws['E1'].font = Font(bold=True, size=13)
    ws['F1'].font = Font(bold=True, size=13)
    ws['G1'].font = Font(bold=True, size=13)
    ws['A1'].fill = PatternFill("solid", start_color="FDD868")
    ws['B1'].fill = PatternFill("solid", start_color="FDD868")
    ws['C1'].fill = PatternFill("solid", start_color="FDD868")
    ws['D1'].fill = PatternFill("solid", start_color="FDD868")
    ws['E1'].fill = PatternFill("solid", start_color="FDD868")
    ws['F1'].fill = PatternFill("solid", start_color="FDD868")
    ws['G1'].fill = PatternFill("solid", start_color="FDD868")
    if not os.path.exists('Αναζητησεις Δρομολογίων'):
        os.makedirs('Αναζητησεις Δρομολογίων')

    file = dates[0]
    if len(dates) == 1:
        wb.save("Αναζητησεις Δρομολογίων/" + file +".xlsx")
    else:
        wb.save("Αναζητησεις Δρομολογίων/" + file+" έως "+dates[-1]+".xlsx")

    print(dates[0],dates[-1])




